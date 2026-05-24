import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ========== 你每次主要改这里 ==========

SOURCE_FILE = r"C:\Users\HUAWEI\Downloads\temp_git_clone\STReasoner_reproduce\00_new_codes\repro_autodl\experiments\stage2_results\experiment1_smarttest\forecasting_prediction.jsonl"
OUTPUT_DIR = r"C:\Users\HUAWEI\Downloads\temp_git_clone\STReasoner_reproduce\00_new_codes\tools\outputs"

# 留空 = 全选所有字段
# 手动指定 = 只展示这些字段，例如 ["id", "question", "answer", "prediction"]
FIELDS = []

# "vertical" = 每个样本一张两列表格：字段 | 内容
# "horizontal" = 所有样本一张横向表格：字段作为表头
LAYOUT = "vertical"

# 最多展示多少条样本
MAX_ROWS = 20

# 每个单元格最多保留多少字符
MAX_CELL_LEN = 300

# 是否同时输出 Excel 文件
OUTPUT_XLSX = True

# =====================================


def load_json_or_jsonl(path: Path):
    suffix = path.suffix.lower()

    if suffix == ".jsonl":
        records = []
        with path.open("r", encoding="utf-8") as f:
            for line_no, line in enumerate(f, start=1):
                line = line.strip()
                if not line:
                    continue

                try:
                    obj = json.loads(line)
                except json.JSONDecodeError as e:
                    raise ValueError(f"第 {line_no} 行不是合法 JSON：{e}")

                if isinstance(obj, dict):
                    records.append(obj)
                else:
                    records.append({"value": obj})

        return records

    if suffix == ".json":
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            return data

        if isinstance(data, dict):
            list_candidates = []

            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    list_candidates.append((key, value))

            if list_candidates:
                list_candidates.sort(key=lambda x: len(x[1]), reverse=True)
                return list_candidates[0][1]

            return [data]

    raise ValueError("只支持 .json 和 .jsonl 文件")


def flatten_dict(d, parent_key="", sep="."):
    result = {}

    if not isinstance(d, dict):
        return {"value": d}

    for key, value in d.items():
        new_key = f"{parent_key}{sep}{key}" if parent_key else str(key)

        if isinstance(value, dict):
            result.update(flatten_dict(value, new_key, sep=sep))
        else:
            result[new_key] = value

    return result


def get_all_fields(flat_records):
    fields = []
    seen = set()

    for record in flat_records:
        for key in record.keys():
            if key not in seen:
                fields.append(key)
                seen.add(key)

    return fields


def format_cell(value):
    if value is None:
        text = ""
    elif isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)

    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = text.replace("|", "\\|")

    if MAX_CELL_LEN is not None and len(text) > MAX_CELL_LEN:
        text = text[:MAX_CELL_LEN] + "..."

    return text


def format_excel_cell(value):
    if value is None:
        text = ""
    elif isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)

    text = text.replace("\r", " ")
    text = text.replace("\n", "\n")

    if MAX_CELL_LEN is not None and len(text) > MAX_CELL_LEN:
        text = text[:MAX_CELL_LEN] + "..."

    return text


def write_horizontal(records, output_path: Path, fields):
    with output_path.open("w", encoding="utf-8") as f:
        f.write("# JSON / JSONL Preview\n\n")
        f.write(f"- Source file: `{SOURCE_FILE}`\n")
        f.write(f"- Total records: {len(records)}\n")
        f.write(f"- Shown records: {min(len(records), MAX_ROWS)}\n")
        f.write(f"- Layout: `{LAYOUT}`\n\n")

        f.write("| " + " | ".join(fields) + " |\n")
        f.write("| " + " | ".join(["---"] * len(fields)) + " |\n")

        for record in records[:MAX_ROWS]:
            flat = flatten_dict(record)
            row = [format_cell(flat.get(field, "")) for field in fields]
            f.write("| " + " | ".join(row) + " |\n")


def write_vertical(records, output_path: Path, fields):
    with output_path.open("w", encoding="utf-8") as f:
        f.write("# JSON / JSONL Preview\n\n")
        f.write(f"- Source file: `{SOURCE_FILE}`\n")
        f.write(f"- Total records: {len(records)}\n")
        f.write(f"- Shown records: {min(len(records), MAX_ROWS)}\n")
        f.write(f"- Layout: `{LAYOUT}`\n\n")

        for i, record in enumerate(records[:MAX_ROWS], start=1):
            flat = flatten_dict(record)

            f.write(f"## Sample {i}\n\n")
            f.write("| 字段 | 内容 |\n")
            f.write("|---|---|\n")

            for field in fields:
                value = format_cell(flat.get(field, ""))
                f.write(f"| `{field}` | {value} |\n")

            f.write("\n---\n\n")


def write_xlsx(records, output_path: Path, fields):
    wb = Workbook()
    ws = wb.active
    ws.title = "preview"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    if LAYOUT == "horizontal":
        # 第一行：字段名
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_alignment

        # 后面每行：一个样本
        for row_idx, record in enumerate(records[:MAX_ROWS], start=2):
            flat = flatten_dict(record)
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=format_excel_cell(flat.get(field, ""))
                )
                cell.alignment = wrap_alignment

        ws.freeze_panes = "A2"

        # 设置列宽
        for col_idx, field in enumerate(fields, start=1):
            letter = get_column_letter(col_idx)
            if field.lower() in ["id", "index", "score", "correct", "label"]:
                ws.column_dimensions[letter].width = 14
            else:
                ws.column_dimensions[letter].width = 35

    elif LAYOUT == "vertical":
        row_idx = 1

        for sample_idx, record in enumerate(records[:MAX_ROWS], start=1):
            flat = flatten_dict(record)

            # Sample 标题行
            title_cell = ws.cell(row=row_idx, column=1, value=f"Sample {sample_idx}")
            title_cell.font = Font(bold=True)
            title_cell.fill = PatternFill("solid", fgColor="EDEDED")
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            row_idx += 1

            # 表头
            field_cell = ws.cell(row=row_idx, column=1, value="字段")
            value_cell = ws.cell(row=row_idx, column=2, value="内容")

            for cell in [field_cell, value_cell]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = wrap_alignment

            row_idx += 1

            # 内容行
            for field in fields:
                key_cell = ws.cell(row=row_idx, column=1, value=field)
                value_cell = ws.cell(
                    row=row_idx,
                    column=2,
                    value=format_excel_cell(flat.get(field, ""))
                )

                key_cell.font = Font(bold=True)
                key_cell.alignment = wrap_alignment
                value_cell.alignment = wrap_alignment

                row_idx += 1

            # 样本之间空一行
            row_idx += 1

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 80

    else:
        raise ValueError('LAYOUT 只能是 "horizontal" 或 "vertical"')

    # 行高稍微放大，方便看长文本
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 45

    wb.save(output_path)


def main():
    source_path = Path(SOURCE_FILE)
    output_dir = Path(OUTPUT_DIR)

    if not source_path.exists():
        raise FileNotFoundError(f"找不到源文件：{source_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    records = load_json_or_jsonl(source_path)

    if not records:
        raise ValueError("文件里没有读到任何记录")

    flat_records = [flatten_dict(record) for record in records[:MAX_ROWS]]

    if FIELDS:
        fields = FIELDS
    else:
        fields = get_all_fields(flat_records)

    md_output_path = output_dir / f"{source_path.stem}_preview.md"
    xlsx_output_path = output_dir / f"{source_path.stem}_preview.xlsx"

    if LAYOUT == "horizontal":
        write_horizontal(records, md_output_path, fields)
    elif LAYOUT == "vertical":
        write_vertical(records, md_output_path, fields)
    else:
        raise ValueError('LAYOUT 只能是 "horizontal" 或 "vertical"')

    print("MD 生成成功：")
    print(md_output_path)

    if OUTPUT_XLSX:
        write_xlsx(records, xlsx_output_path, fields)
        print("Excel 生成成功：")
        print(xlsx_output_path)


if __name__ == "__main__":
    main()